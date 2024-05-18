import os
import time
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.edge.options import Options as EdgeOptions

def extract_from_ninth_digit(text):
    # Split the input text into lines
    lines = text.splitlines()

    # Initialize an empty list to store the extracted portions
    extracted_portions = []

    # Iterate through each line
    for line in lines:
        # Check if the line has at least 9 characters
        if len(line) >= 9:
            # Extract from the 9th digit to the end
            extracted_portion = line[9:]
            extracted_portions.append(extracted_portion)
        else:
            # If the line is shorter than 9 characters, skip it
            extracted_portions.append("")

    return extracted_portions


def findValueInColumn(file_path, column_letter, search_value):
    matchingRow = 0
    try:
        workbook = load_workbook(file_path)
        sheet = workbook.active
        i = 2

        column_index = ord(column_letter.upper()) - ord('A') + 1

        for row in sheet.iter_rows(min_row=2, min_col=column_index, max_col=column_index, max_row=sheet.max_row):
            cell_value = row[0].value
    
            if cell_value == search_value:
                matchingRow = i
            i = i + 1
        
        if matchingRow != 0:
            return matchingRow
        else: 
            return None

    except FileNotFoundError:
        print(f"File '{file_path}' not found.")
        return []

def count_non_empty_rows(sheet):
    try:
        non_empty_rows = sum(1 for row in sheet.iter_rows() if any(cell.value for cell in row))
        return non_empty_rows
    except FileNotFoundError:
        return 0


def remove_first_digit_if_starts_with_9(number):
    # Convert the input to a string for easier manipulation
    number_str = str(number)

    # Check if the number starts with 9
    if number_str.startswith('9'):
        # Remove the first digit
        modified_number_str = number_str[1:]
        return int(modified_number_str) if number_str.isdigit() else modified_number_str
    else:
        return number

def get_last_characters(text):
    lines = text.splitlines()  # Split the text into individual lines
    last_chars = [line.strip()[-1] for line in text.splitlines()]
    return "".join(last_chars)

def array_to_lines(arr):
    return '\n'.join(str(item) for item in arr)


excelName = input("Ingrese el nombre del Excel con el cual trabajara el Script: \n")#NAME OF THE EXCEL FILE
excelFile = excelName + ".xlsx" #ADD EXTENSION FILE
#excel_path = os.path.join(os.path.dirname(__file__), excelFile)

workbook = load_workbook(filename = str(excelFile))
sheet = workbook.active

option = EdgeOptions()
option.add_argument("start-maximized")
driver = webdriver.Edge(options = option)

timeOut = 25

EIDArray = []
authArray = []
i = 1



for row in sheet.iter_rows(min_row=2, min_col=1, max_row=(sheet.max_row), max_col=1):
    i = i + 1
    if sheet["B" + str(i)].value != None:
        driver.get("https://directory.accenture.com//UserStatus")
        os.system("CLS")
        WebDriverWait(driver, timeOut).until(EC.title_is("Accenture User Status"))

        try:
            WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[3]/div[2]/form/div[3]/table/tbody/tr[3]/td[2]/table/tbody/tr/td[1]/input')))
            personalNumber = driver.find_element(By.XPATH, '/html/body/div[3]/div[2]/form/div[3]/table/tbody/tr[3]/td[2]/table/tbody/tr/td[1]/input')
            personalNumber.clear()
            personalNumber.send_keys(sheet["B" + str(i)].value)
        except TimeoutException:
            print ("Web tool failed, run script again...")


        try:
            WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[3]/div[2]/form/div[3]/table/tbody/tr[6]/td[2]/input[1]')))
            getInfo = driver.find_element(By.XPATH, '/html/body/div[3]/div[2]/form/div[3]/table/tbody/tr[6]/td[2]/input[1]')
            getInfo.click()
        except TimeoutException:
            print ("Web tool failed, run script again...")
        

        try:
            WebDriverWait(driver, timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[3]/div[2]/form/div[3]/table/tbody/tr[15]/td/div[1]/div[1]/table/tbody/tr/td[2]/div')))
            EIDExcel = driver.find_element(By.XPATH, '/html/body/div[3]/div[2]/form/div[3]/table/tbody/tr[15]/td/div[1]/div[1]/table/tbody/tr/td[2]/div')
            sheet["C" + str(i)] = EIDExcel.text
        except TimeoutException:
            if driver.find_element(By.XPATH, '/html/body/div[3]/div[2]/form/div[3]/table/tbody/tr[12]/td/span'):
                sheet["C" + str(i)] = 'Wrong PN'
            print ("Web tool, run script again...")
            
        
    if sheet["C" + str(i)].value != 'Wrong PN':
        authField = str(sheet["C" + str(i)].value) + "," + "+549" + str(remove_first_digit_if_starts_with_9(sheet["W" + str(i)].value))

        EIDArray.append(sheet["C" + str(i)].value)
        authArray.append(authField)
    else:
        print("El archivo excel no tiene NJ")
workbook.save(excelFile)
driver.get("https://directory.accenture.com/ResetPassword/BulkReset")

time.sleep(3)
os.system('cls')




try:
    WebDriverWait(driver,timeOut).until(EC.presence_of_element_located((By.XPATH, "/html/body/article/div[4]/div[1]/div[1]/div[2]/div[3]")))
    njResetButton = driver.find_element(By.XPATH, "/html/body/article/div[4]/div[1]/div[1]/div[2]/div[3]")
    njResetButton.click()
except TimeoutException:
    print ("Web tool failed, run script again...")

time.sleep(2)

try:
    WebDriverWait(driver,timeOut).until(EC.presence_of_element_located(((By.XPATH, "/html/body/article/div[4]/div/div/form/div/div[3]/div[2]/div[1]/div[3]/textarea"))))
    userArea = driver.find_element(By.XPATH, "/html/body/article/div[4]/div/div/form/div/div[3]/div[2]/div[1]/div[3]/textarea")
    userArea.send_keys(array_to_lines(EIDArray))
except TimeoutException:
    print ("Web tool failed, run script again...")
    


try:
    WebDriverWait(driver,timeOut).until(EC.presence_of_element_located((By.XPATH, "/html/body/article/div[4]/div/div/form/div/div[3]/div[2]/div[2]/div[3]/input[1]")))
    dirButton = driver.find_element(By.XPATH, "/html/body/article/div[4]/div/div/form/div/div[3]/div[2]/div[2]/div[3]/input[1]")
    dirButton.click()
except TimeoutException:
    print ("Web tool failed, run script again...")



try:
    WebDriverWait(driver,timeOut).until(EC.presence_of_element_located((By.XPATH, "/html/body/article/div[4]/div/div/form/div/div[4]/div/input[1]")))
    resetPaswdButton = driver.find_element(By.XPATH, "/html/body/article/div[4]/div/div/form/div/div[4]/div/input[1]")
    resetPaswdButton.click()
    print("Resseting passwords")
except TimeoutException:
    print ("Web tool failed, run script again...")

i = 2

for row in sheet.iter_rows(min_row=2, min_col=1, max_row=(len(EIDArray)) + 1, max_col=1):
    if sheet["C" + str(i)] != 'Wrong PN':
        try:
            print(i)
            WebDriverWait(driver,120).until(EC.presence_of_element_located((By.XPATH, "/html/body/article/div[4]/div/div/form/div/div[5]/div/div[2]/table/tbody/tr["+str(i)+"]/td[2]")))
            EID = driver.find_element(By.XPATH, "/html/body/article/div[4]/div/div/form/div/div[5]/div/div[2]/table/tbody/tr["+str(i)+"]/td[1]")

            matchingRow = findValueInColumn(file_path=str(excelFile), column_letter='C',search_value= EID.text)
            passwordField = driver.find_element(By.XPATH, "/html/body/article/div[4]/div/div/form/div/div[5]/div/div[2]/table/tbody/tr["+str(i)+"]/td[2]")

            sheet["AJ" + str(matchingRow)] = passwordField.text
            try:
                WebDriverWait(driver,timeOut).until(EC.presence_of_element_located((By.XPATH, "/html/body/article/div[4]/div/div/form/div/div[5]/div/div[2]/table/tbody/tr["+str(i)+"]/td[4]")))
                resultPassword = driver.find_element(By.XPATH, "/html/body/article/div[4]/div/div/form/div/div[5]/div/div[2]/table/tbody/tr["+str(i)+"]/td[4]")
                sheet["AL" + str(matchingRow)] = resultPassword.text[:12]
            except TimeoutException:
                print ("Web tool failed, run script again...")
            

        except TimeoutException:
            print ("Web tool failed, run script again...")
        i = i + 1
        workbook.save(excelFile)

driver.get("https://myuserauth.accenture.com")
os.system('cls')
time.sleep(5)
try:
    WebDriverWait(driver,timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/app-root/body/div/app-tapservices/mat-tab-group/mat-tab-header/div/div/div/div[4]')))
    phoneTab = driver.find_element(By.XPATH, '/html/body/app-root/body/div/app-tapservices/mat-tab-group/mat-tab-header/div/div/div/div[4]')
    phoneTab.click()
except TimeoutException:
    print("AUTH SITE FAILED. RUN SCRIPT AGAIN")
    print("possible user account syntax error or web issue")


try:
    WebDriverWait(driver,timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/app-root/body/div/app-tapservices/mat-tab-group/div/mat-tab-body[4]/div/app-bulk-add-phone-number/div/form/div[2]/mat-form-field/div[1]/div/div[2]/textarea')))
    eidField = driver.find_element(By.XPATH, '/html/body/app-root/body/div/app-tapservices/mat-tab-group/div/mat-tab-body[4]/div/app-bulk-add-phone-number/div/form/div[2]/mat-form-field/div[1]/div/div[2]/textarea')
    eidField.send_keys(array_to_lines(authArray))
except TimeoutException:
    print("AUTH SITE FAILED. RUN SCRIPT AGAIN")
    print("possible user account syntax error or web issue")


try:
    WebDriverWait(driver,timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/app-root/body/div/app-tapservices/mat-tab-group/div/mat-tab-body[4]/div/app-bulk-add-phone-number/div/form/div[4]/div[2]/div[1]/mat-radio-button/div/div/input')))
    NJButton = driver.find_element(By.XPATH, '/html/body/app-root/body/div/app-tapservices/mat-tab-group/div/mat-tab-body[4]/div/app-bulk-add-phone-number/div/form/div[4]/div[2]/div[1]/mat-radio-button/div/div/input')
    NJButton.click()
except TimeoutException:
    print("AUTH SITE FAILED. RUN SCRIPT AGAIN")
    print("possible user account syntax error or web issue")

try:
    WebDriverWait(driver,timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/div[3]/div/div[1]')))
    hideButton = driver.find_element(By.XPATH, '/html/body/div[3]/div/div[1]')
    hideButton.click()
except TimeoutException:
    print("AUTH SITE FAILED. RUN SCRIPT AGAIN")
    print("possible user account syntax error or web issue")

try:
    WebDriverWait(driver,timeOut).until(EC.presence_of_element_located((By.XPATH, '/html/body/app-root/body/div/app-tapservices/mat-tab-group/div/mat-tab-body[4]/div/app-bulk-add-phone-number/div/div/button')))
    submitButton = driver.find_element(By.XPATH, '/html/body/app-root/body/div/app-tapservices/mat-tab-group/div/mat-tab-body[4]/div/app-bulk-add-phone-number/div/div/button')
    submitButton.click()
    print("Adding phones to MFA")
except TimeoutException:
    print("AUTH SITE FAILED. RUN SCRIPT AGAIN")
    print("possible user account syntax error or web issue")


try:
    WebDriverWait(driver,120).until(EC.presence_of_element_located((By.XPATH, '/html/body/app-root/body/div/app-tapservices/mat-tab-group/div/mat-tab-body[4]/div/app-bulk-add-phone-number/div/div[2]/mat-form-field/div[1]/div/div[2]/textarea')))
    resultInfo = driver.find_element(By.XPATH, '/html/body/app-root/body/div/app-tapservices/mat-tab-group/div/mat-tab-body[4]/div/app-bulk-add-phone-number/div/div[2]/mat-form-field/div[1]/div/div[2]/textarea')
    result = resultInfo.text
    successUserList = extract_from_ninth_digit(result)

    for user in successUserList:
        matchingRow = findValueInColumn(file_path=str(excelFile),column_letter='C',search_value=user)
        sheet['AK'+str(matchingRow)] = 'Successful'

except TimeoutException:
    print("TAP SITE FAILED. RUN SCRIPT AGAIN")
    print("possible user account syntax error or web issue")
workbook.save(excelFile)
